from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def xlsx_to_dict( fil_path ):
    xlsx_file = Path( fil_path )
    wb_obj = load_workbook(xlsx_file)
    sheet = wb_obj.active
    max_row = sheet.max_row
    max_column = sheet.max_column

    skills = {}
    for column_number in range(1, max_column + 1):
        if column_number % 2 != 0:
            skills_names = []
            skills_levels = []

        column_letter = get_column_letter( column_number )
        for row in range(1, max_row + 1):
            if sheet[f'{ column_letter }{ row }'].value is not None:
                if column_number % 2 != 0:
                    skills_names.append(sheet[f'{ column_letter }{ row }'].value)
                else:
                    skills_levels.append(sheet[f'{ column_letter }{ row }'].value)
            else:
                break

        if column_number % 2 == 0:
            skill_section = skills_names[0]
            skills_names.pop(0)
            skills_levels.pop(0)
            skills_section = []
            for name, level in zip(skills_names, skills_levels):
                skills_section.append(
                    {
                        'name': name,
                        'level': int(level)
                    }
                )

            skills[skill_section] = skills_section

    return skills
